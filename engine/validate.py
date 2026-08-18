#!/usr/bin/env python3
"""Prove the demo is safe to run in front of a client -- all three datasets of it.

Everything here answers one question: will the screen show what we expect it to show?
A demo is only rehearsable if the answer is fixed, so this checks the engine reproduces
each generator's ground truth exactly, that every control passes, and that the result
actually makes the case -- a saving with a real spread of good and bad lanes, not a
uniformly positive answer nobody believes.

Three datasets rather than one, because the demo's claim is not "it works on our file"
but "it reads whatever file you have and tells you what it can and cannot price". That
claim is only testable against files of different quality, and the expectations below
are deliberately *different* per scenario: Northgate derives almost everything, Meritt
derives one leg out of eight. A suite that expected the same board from both would be
asserting the opposite of what the product does.

It also round-trips the mapping files, because "upload these next time and the review
step is instant" is a claim we make out loud and it has to be true.

    python3 engine/validate.py
    python3 engine/validate.py --scenario meritt

Exits non-zero on any failure, so it can gate a commit.
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd

import config as C
import costing as costing_mod
import rates as rates_mod
import reconcile as reconcile_mod
import resolve as resolve_mod
import run as run_mod
import sourcing as sourcing_mod

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"


# --------------------------------------------------------------------------------------
# The rehearsal lock.
#
# One entry per dataset. These are engine properties rather than data properties -- how
# many decisions escalate, what state each leg lands in -- so they live here and not in
# the generator's manifest. A number moving between rehearsal and the meeting is a demo
# you cannot script, which is the whole reason this file exists.
#
# The board expectations are the interesting ones, and they are stated for the board with
# NO cards uploaded, because that is the state the demo opens on and the only one that
# depends purely on the client's own file.
# --------------------------------------------------------------------------------------
SCENARIOS = {
    "northgate": {
        "queue": 3,
        "queue_rules": {"not_a_place", "operator_multi_city",
                        "supplier_office_not_origin"},
        # Itemised billing, so every leg they already buy prices itself, and the
        # warehouse-to-port leg has code 1602 to reach for -- the analogue. What is left
        # unpriced is the consolidation service, which no invoice can evidence and which
        # the engine will not price from a benchmark of ours.
        "bare_board": {"derived": 5, "analogue": 1, "unpriced": 3},
        "bare_unpriced": ["warehouse_handling", "warehouse_inbound",
                          "warehouse_storage"],
        "blocked_bundles": [],
        "ftl": "adopted",
        # The strong case. They buy in small lots from many suppliers and book each one on
        # its own, so their boxes leave 57% full and there is a great deal of air to take
        # out. It clears the bar under every setting -- worst case still +10%.
        "case": "strong",
        "saving_pct_band": (0.14, 0.21),
        "pooling": "better",
    },
    "calderwood": {
        "queue": 3,
        "queue_rules": {"not_a_place", "operator_multi_city",
                        "supplier_office_not_origin"},
        # Origin bundled into 1630, so both origin legs are unpriceable from the file.
        # No code 1602 either, so warehouse-to-port has no analogue to fall back on and
        # waits for the quote like the rest of the service.
        "bare_board": {"derived": 3, "unpriced": 6},
        "bare_unpriced": ["origin_collection", "origin_terminal",
                          "warehouse_handling", "warehouse_inbound",
                          "warehouse_storage", "warehouse_to_port"],
        "blocked_bundles": ["1630"],
        "ftl": "rejected",
        # The borderline case. Bigger consignments and tighter bookings, so the saving is
        # real and small, and the settings decide whether it survives: letting sites share
        # a box takes it from 6% to 11%, and a three-day dwell takes it to 3%.
        "case": "borderline",
        "saving_pct_band": (0.04, 0.09),
        "pooling": "better",
    },
    "meritt": {
        # Deliberately not
        # all four rules. This world's multi-location operator is two spellings of one
        # town, which the engine settles on its own, so nothing escalates for it.
        "queue": 3,
        "queue_rules": {"not_a_place", "operator_multi_city",
                        "supplier_office_not_origin"},
        # Both bundles. Exactly one leg still derives, and seven of the eight wait for a
        # file -- which is the honest reading of a file this coarse.
        "bare_board": {"derived": 1, "unpriced": 8},
        "bare_unpriced": ["destination_delivery", "ocean_freight",
                          "origin_collection", "origin_terminal",
                          "warehouse_handling", "warehouse_inbound",
                          "warehouse_storage", "warehouse_to_port"],
        "blocked_bundles": ["1300", "1630"],
        "ftl": "absent",
        # The marginal case. Container-sized ordering leaves much less air than the other
        # worlds, but a few fragmented site lanes still clear the same commercial rule.
        # Three lanes run, four are left alone, and the saving stays in low single digits.
        "case": "marginal",
        "saving_pct_band": (0.025, 0.05),
        "pooling": "worse",
    },
}


def close(a, b, tol=0.05):
    return abs(float(a) - float(b)) <= tol


def lanes_of(result):
    return list(result["lanes"]["verdict"])


class Checks:
    """A scenario's results, so several scenarios can be run and summed."""

    def __init__(self, label):
        self.label = label
        self.rows = []

    def __call__(self, name, passed, detail=""):
        self.rows.append((name, bool(passed), detail))
        print(f"  {'PASS' if passed else 'FAIL'}  {name}"
              + (f"\n          {detail}" if detail and not passed else ""))
        return bool(passed)

    @property
    def failed(self):
        return [n for n, ok, _ in self.rows if not ok]


def validate_scenario(scenario_id, spec):
    d = DATA / scenario_id
    charge_lines = str(d / "charge_lines_raw.csv")
    manifest = json.loads((d / "ground_truth_manifest.json").read_text())
    # No site list. The app no longer asks for one and the generator no longer writes
    # one: the engine reads the client's warehouses out of their own delivery addresses.
    # A list is still *accepted*, and the checks below build one on the fly to prove it
    # still overrides -- which is a better test than shipping a file, because it spells a
    # town differently on purpose.
    site_list = None
    # The reference pack is the demo path: one upload answering everything the client
    # already knows the price of.
    rate_card = pd.read_csv(d / "reference_pack.csv")
    ftl_path = d / "ftl_rate_card.csv"

    check = Checks(scenario_id)
    print(f"\n{'=' * 74}\n  {manifest['client']}  ({scenario_id})\n{'=' * 74}")

    print("\nThe engine reads the file the generator wrote")
    ingested, resolution = run_mod.preview_queue(
        charge_lines, rate_card=rate_card, site_list=site_list)
    st = ingested.stats
    check("charge lines match the manifest",
          st["charge_lines"] == manifest["charge_lines"],
          f"{st['charge_lines']} vs {manifest['charge_lines']}")
    check("shipments match", st["shipments"] == manifest["shipments"],
          f"{st['shipments']} vs {manifest['shipments']}")
    check("shipment groups match",
          st["shipment_groups"] == manifest["shipment_groups"],
          f"{st['shipment_groups']} vs {manifest['shipment_groups']}")
    check("invoiced total matches to the cent",
          close(st["total_invoiced_usd"], manifest["total_invoiced_usd"]),
          f"{st['total_invoiced_usd']} vs {manifest['total_invoiced_usd']}")
    check("in-scope groups match",
          st["in_scope_groups"] == manifest["in_scope"]["shipment_groups"],
          f"{st['in_scope_groups']} vs {manifest['in_scope']['shipment_groups']}")
    check("containers rebuilt from master bills match",
          st["containers_today"] == manifest["in_scope"]["containers_today"],
          f"{st['containers_today']} vs {manifest['in_scope']['containers_today']}")
    check("in-scope pallets match",
          st["in_scope_pallets"] == manifest["in_scope"]["pallets"],
          f"{st['in_scope_pallets']} vs {manifest['in_scope']['pallets']}")

    print("\nThe mess is present and was handled, not skipped")
    mess = manifest["mess"]
    sentinel = next(f for f in ingested.findings if f.rule == "repeated_invoice_total")
    check("the '-related-' sentinel was found on every marked row",
          sentinel.rows == mess["related_sentinel_rows"],
          f"{sentinel.rows} vs {mess['related_sentinel_rows']}")
    check("reading the repeated column would have overstated spend",
          sentinel.dollars > 0,
          f"overstatement of ${sentinel.dollars:,.2f}")
    unknown = ingested.lines[ingested.lines["code_unrecognised"]]
    check("every generated charge code is in the register",
          len(unknown) == 0
          and mess["unrecognised_codes"] == []
          and sorted(unknown["code"].unique()) == mess["unrecognised_codes"]
          and close(float(unknown["line_usd"].sum()), mess["unrecognised_code_usd"], 0.5),
          f"{sorted(unknown['code'].unique())} / ${float(unknown['line_usd'].sum()):,.2f}")
    unknown_probe = ingested.lines.copy()
    unknown_probe.loc[unknown_probe.index[0], "code_unrecognised"] = True
    check("a genuinely unknown code blocks a usable result",
          not reconcile_mod.charge_code_control(unknown_probe).passed)

    print("\nThe review queue is the size we rehearsed")
    rules = {q.rule for q in resolution.queue}
    check(f"exactly {spec['queue']} decisions escalate",
          len(resolution.queue) == spec["queue"],
          f"{len(resolution.queue)} escalated: {sorted(rules)}")
    check("the expected decision types are represented",
          rules == spec["queue_rules"],
          f"missing {sorted(spec['queue_rules'] - rules)}, "
          f"extra {sorted(rules - spec['queue_rules'])}")
    check("every escalation carries a proposal and its evidence",
          all(q.proposal and q.evidence and q.options for q in resolution.queue))
    uids = [q.uid for q in resolution.queue]
    check("every escalation has a unique identity",
          len(set(uids)) == len(uids),
          f"{len(uids) - len(set(uids))} decisions share an identity — answers would "
          "cross-apply and the interface would reject the duplicate widget key")
    # Two claims to test, because between them they are the reason the file is gone.
    #
    # First, that the warehouses really are read out of the client's own delivery
    # addresses. Second, that a list is still an aid where one exists -- so it is built
    # here rather than read, and built badly on purpose: it names two of the resolved
    # towns with a country the file never gave, and spells one of them differently.
    from_data = [r for r in resolution.sites.to_dict("records")
                 if "Identified from your own file" in str(r.get("Note", ""))]
    check("the warehouses are read out of the client's own delivery addresses",
          len(from_data) >= 3, f"{len(from_data)} sites identified from the file")

    towns = sorted({str(n) for n in resolution.sites["Site_Name"]})
    invented = pd.DataFrame([
        {"Site_ID": f"XX_{towns[0].upper()}", "Site_Name": f"{towns[0]}, XX",
         "Site_Country": "Elsewhere", "Served_By_Port": ""},
        # The same town with a letter dropped, which is what a commercial team's
        # spreadsheet actually looks like next to an operations export.
        {"Site_ID": f"XX_{towns[1].upper()}", "Site_Name": f"{towns[1][:-1]}, XX",
         "Site_Country": "Elsewhere", "Served_By_Port": ""},
    ])
    _, with_list = run_mod.preview_queue(charge_lines, rate_card=rate_card,
                                         site_list=invented)
    listed = set(with_list.sites["Site_Name"])
    check("a supplied site list still overrides what the file said",
          f"{towns[0]}, XX" in listed,
          f"{sorted(listed)[:4]} — expected '{towns[0]}, XX' to win")
    check("and a town it spells differently is matched, not duplicated",
          f"{towns[1][:-1]}, XX" in listed and towns[1] not in listed,
          f"both '{towns[1]}' and '{towns[1][:-1]}, XX' resolved" )
    check("supplying one does not lengthen the review",
          len(with_list.queue) <= len(resolution.queue),
          f"{len(with_list.queue)} with a list against {len(resolution.queue)} without")

    check("most mappings resolve without a human",
          resolution.stats["auto_resolved"] >= 4 * len(resolution.queue),
          f"{resolution.stats['auto_resolved']} auto vs {len(resolution.queue)} escalated")
    # A region proposed for a supplier has to be one this client actually collects from.
    # The engine's gazetteer spans every origin it knows, and offering a Vietnamese
    # region to a German importer buying out of Qingdao is the kind of detail that loses
    # a room.
    regions_in_file = set(resolution.suppliers["Rate_Region"].dropna().unique())
    supplier_items = [q for q in resolution.queue if q.kind == "supplier"]
    # Checked on the values the answers carry, not on their wording: the options are
    # sentences a reviewer reads, and the regions behind them are what the model applies.
    check("supplier region options are drawn from this client's own origins",
          all(set(q.values.values()) <= regions_in_file for q in supplier_items),
          "; ".join(f"{q.key}: {sorted(set(q.values.values()) - regions_in_file)}"
                    for q in supplier_items
                    if not set(q.values.values()) <= regions_in_file))
    check("every option a reviewer can pick carries a value the model understands",
          all(set(q.options) - set(q.values) <= {resolve_mod.EXCLUDE,
                                                 "Different places — keep them separate"}
              for q in resolution.queue),
          "; ".join(f"{q.key}: {sorted(set(q.options) - set(q.values))}"
                    for q in resolution.queue if set(q.options) - set(q.values)
                    - {resolve_mod.EXCLUDE, "Different places — keep them separate"}))

    print("\nThe sourcing board reads the file, with no cards given")
    bare_res = resolve_mod.resolve(ingested.shipments, site_list=site_list)
    bare_ship = resolve_mod.apply_resolution(ingested.shipments, bare_res)
    bare_plan = sourcing_mod.plan(ingested.lines, bare_ship, {},
                                  containers_today=int(st["containers_today"]))
    bare_board = sourcing_mod.summarise(bare_plan)
    check("every component resolves to a state",
          len(bare_plan) == len(C.COST_COMPONENTS)
          and all(cs.state in sourcing_mod.PRICED_STATES | {sourcing_mod.UNPRICED}
                  for cs in bare_plan.values()),
          str(bare_board["by_state"]))
    check("the board is the shape this file implies",
          bare_board["by_state"] == spec["bare_board"],
          f"{bare_board['by_state']} vs expected {spec['bare_board']}")
    check("the legs that cannot be priced are the ones we expect",
          bare_board["unpriced"] == spec["bare_unpriced"],
          f"{bare_board['unpriced']} vs {spec['bare_unpriced']}")
    check("bundled charge codes are named as the reason, not left as an absence",
          bare_board["blocked_by_bundles"] == spec["blocked_bundles"],
          f"{bare_board['blocked_by_bundles']} vs {spec['blocked_bundles']}")
    for cs in bare_plan.values():
        if cs.state == sourcing_mod.UNPRICED:
            check(f"'{cs.label}' says why it cannot be priced",
                  bool(cs.rejected) and all(r["why"] for r in cs.rejected))
    check("an analogue never reads as high confidence",
          all(cs.state != sourcing_mod.ANALOGUE or cs.caveat
              for cs in bare_plan.values()),
          "an analogue with no caveat would present adjacent work as a clean derivation")

    print("\nThe reference pack closes every gap")
    packed = sourcing_mod.plan(ingested.lines, bare_ship,
                               rates_mod.index_rate_card(rate_card),
                               containers_today=int(st["containers_today"]))
    packed_board = sourcing_mod.summarise(packed)
    check("nothing is left unpriced once the pack is supplied",
          not packed_board["unpriced"], str(packed_board["unpriced"]))
    check("every leg is priced", packed_board["priced"] == len(C.COST_COMPONENTS),
          f"{packed_board['priced']} of {len(C.COST_COMPONENTS)}")

    print("\nThe full run, accepting the engine's proposals")
    answers = {q.uid: q.proposal for q in resolution.queue}
    result = run_mod.run(charge_lines, rate_card=rate_card, site_list=site_list,
                         answers=answers)
    s, cs_sum = result["summary"], result["controls_summary"]

    check("every control passes", cs_sum["all_passed"],
          "failed: " + ", ".join(cs_sum["failures"]))
    check("the ledger is the modelled cost to the cent",
          close(float(result["ledger"]["usd"].sum()), s["future"]["Total"], 0.01))
    check("no dollars fall out between the file and the pools",
          close(float(result["lines"]["line_usd"].sum()),
                float(result["shipments"][C.COST_POOLS].sum().sum()), 0.05))
    check("no rate is invented",
          s["provenance"]["rates_invented"] == 0
          and result["rate_table"].stats["gaps"] == 0,
          f"{result['rate_table'].stats['gaps']} unpriced legs")
    check("with the service quoted, every modelled dollar is evidenced",
          abs(s["provenance"]["evidenced_share"] - 1.0) < 0.001,
          f"{s['provenance']['evidenced_share']:.2%} evidenced")
    derived = result["rates"][result["rates"]["source"].eq("DERIVED_FROM_INVOICES")]
    check("every derived rate traces to at least one invoice",
          bool(len(derived)) and bool((derived["population_invoices"] > 0).all()),
          f"{len(derived)} derived rates")
    # The consolidation quote has to actually reach the arithmetic. Reading our own
    # config while telling the client we applied their forwarder's number is the one
    # claim in the model the file cannot support, and nothing downstream would notice.
    quote_index = rates_mod.index_rate_card(rate_card)
    drayage = result["rate_table"].get("SERVICE", "cfs_drayage")
    quoted_drayage = quote_index.get(("CONSOLIDATION", "", "", "CFS_DRAYAGE"))
    check("the uploaded consolidation quote is the number the model used",
          quoted_drayage is not None
          and close(drayage["rate_usd"], quoted_drayage["rate_usd"], 0.01),
          f"model ${drayage['rate_usd']:,.2f} against a quoted "
          f"${(quoted_drayage or {}).get('rate_usd', 0):,.2f}")

    # --- the case each dataset is here to make ----------------------------------
    #
    # Not merely "the saving is positive". Each dataset is tuned to a distinct strength,
    # and every one must still contain both a useful opportunity and lanes the model turns
    # down. That keeps every demo productive without making any one uniformly flattering.
    print(f"\nThe result makes the case this dataset is here to make: {spec['case']}")
    lo, hi = spec["saving_pct_band"]
    check(f"the saving lands in the {lo:+.0%} to {hi:+.0%} band this dataset is tuned to",
          lo <= s["saving_pct"] <= hi,
          f"{s['saving_pct']:+.1%} (${s['saving_usd']:,.0f})")
    check("the saving is positive", s["saving_usd"] > 0, f"${s['saving_usd']:,.0f}")
    check("containers fall", s["containers_saved"] > 0,
          f"{s['containers_today']} -> {s['containers_future']}")
    check("containers are fuller than they are today",
          s["mean_fill_cbm"] > s["today_fill_cbm"],
          f"{s['today_fill_cbm']} -> {s['mean_fill_cbm']} CBM")
    check("lanes disagree with each other", len(set(lanes_of(result))) >= 2,
          f"verdicts present: {sorted(set(lanes_of(result)))}")
    worth = sum(1 for v in lanes_of(result) if v == "Consolidate")
    check("some lanes are clearly worth doing", worth >= 3,
          f"{worth} of {len(lanes_of(result))} lanes")
    check("origin cost rises only where the final plan actually uses the warehouse",
          ((s["future"]["Origin CFS"] > 0 and s["today"]["Origin CFS"] == 0)
           if s["consolidated_lanes"] else
           close(s["future"]["Origin CFS"], s["today"]["Origin CFS"], 0.05)),
          f"{s['consolidated_lanes']} adopted lanes, "
          f"${s['future']['Origin CFS']:,.0f} future origin CFS")

    # The gate. Asserted on every scenario and on both pool keys, because it is the one
    # promise the model makes about lanes it cannot help: it will not touch them. A single
    # lane gaining a box would mean the plan is buying warehouse handling to make a lane
    # worse, and no lane verdict wording could rescue that.
    print("\nThe plan never adds a container to a lane")
    lanes = result["lanes"]
    gained = lanes[lanes["containers_saved"] < 0]
    check("no lane ships in more containers than it does today", len(gained) == 0,
          "; ".join(f"{r['cfs']}->{r['site']} {int(r['containers_today'])}"
                    f"->{int(r['containers_future'])}" for _, r in gained.iterrows()))
    declined = lanes[~lanes["consolidated"]]
    # To the cent, allowing a cent a container: the cost of a declined box is its own
    # invoice pro-rated over its pallets and rounded where it is posted, so a 42-box lane
    # can land four cents from its invoiced total. Anything larger than that is a real
    # difference and means something was re-priced.
    off = [(r["site"], r["saving_usd"]) for _, r in declined.iterrows()
           if r["containers_saved"] != 0
           or abs(r["saving_usd"]) > 0.01 * max(1, r["containers_today"]) + 0.05]
    check("a declined lane is handed back untouched — same boxes, same cost",
          not off, f"{len(declined)} declined lanes; off by: {off}")
    check("and the containers on it are all carried at their own invoices",
          s["declined_containers"] <= s["passthrough_containers"],
          f"{s['declined_containers']} declined of {s['passthrough_containers']} passthrough")
    check("the boxes the plan reports are the boxes the client would book",
          s["containers_future"] == int(result["containers_costed"]["counts_as_boxes"].sum()),
          f"{s['containers_future']}")
    ledger_declined = result["ledger"][
        result["ledger"]["container"].isin(
            result["containers_costed"].loc[
                ~result["containers_costed"]["consolidated"], "container"])]
    check("nothing new is charged to a lane the plan declined",
          set(ledger_declined["rate_id"].unique()) <= {"ACTUAL-INVOICE"},
          f"{sorted(set(ledger_declined['rate_id'].unique()))}")

    # Letting sites share a container is the one setting that changes the answer's shape,
    # so what it does on this file is pinned rather than left to the room to discover. It
    # fills boxes fuller and buys a destination strip for every box carrying two sites,
    # and which of those wins is a property of the file.
    mixed = run_mod.run(charge_lines, rate_card=rate_card, site_list=site_list,
                        answers=answers, overrides={"POOL_KEY": "cfs_pod"})
    ms, mled = mixed["summary"], mixed["ledger"]
    strip = float(mled[mled["rate_id"] == "NEW-DECONSOL"]["usd"].sum())
    mixed_alloc = pd.DataFrame(mixed["allocation"])
    site_counts = mixed_alloc.groupby("container")["site"].nunique()
    consolidated_ids = set(mixed["containers_costed"].loc[
        mixed["containers_costed"]["consolidated"], "container"])
    actual_mixed_boxes = any(i in consolidated_ids for i in site_counts[site_counts > 1].index)
    check("mixed-site boxes buy a destination strip, and rejected candidates do not",
          (strip > 0) == actual_mixed_boxes,
          f"${strip:,.0f} across {ms['containers_future']} final containers")
    check(f"and on this file it makes the answer {spec['pooling']}",
          (ms["saving_usd"] > s["saving_usd"]) == (spec["pooling"] == "better"),
          f"${s['saving_usd']:,.0f} one site per box against ${ms['saving_usd']:,.0f} mixed, "
          f"after ${strip:,.0f} of strip")
    check("the mixed-site route reaches the dashboard with every control passing",
          mixed["controls_summary"]["all_passed"],
          ", ".join(mixed["controls_summary"]["failures"]))

    lt = result["lead_time"]
    warehouse_containers = result["containers_costed"][
        ~result["containers_costed"]["passthrough"]]
    dispatch_total = sum(lt["dispatch_reasons"].values())
    check("warehouse dwell and dispatch reasons cover only warehouse-built containers",
          dispatch_total == len(warehouse_containers) == s["cfs_built_containers"],
          f"{len(warehouse_containers)} warehouse containers, "
          f"{dispatch_total} dispatch reasons, "
          f"{s['cfs_built_containers']} reported as warehouse-built")
    passthrough = result["containers_costed"][result["containers_costed"]["passthrough"]]
    check("untouched shipments never enter the lead-time model",
          all(result["deltas"].loc[
              result["deltas"]["grp_key"].isin(result["passthrough_groups"]),
              "moved"].eq(False)),
          f"{len(passthrough)} untouched containers")
    check("the service impact is conservative, and zero only for an unchanged plan",
          lt["mean_delta_days_conservative"] >= lt["mean_delta_days"]
          and ((lt["mean_delta_days_conservative"] > 0) == bool(s["consolidated_lanes"])),
          f"conservative {lt['mean_delta_days_conservative']} vs raw {lt['mean_delta_days']}")
    check("no cargo breaches the dwell cap", lt["cap_breaches"] == 0)
    check("every container sails on a departure the lane actually ran",
          lt["containers_beyond_last_sailing"] == 0,
          f"{lt['containers_beyond_last_sailing']} containers past the last recorded sailing")

    # The lead-time panels are the part of this answer an operations audience reads
    # first, and every headline on them is arithmetic off these numbers. So the
    # arithmetic is checked rather than the wording: a panel that does not add up is a
    # sentence on the screen that is not true.
    print("\nThe lead-time panels add up to the run behind them")
    panels = result["lead_time_origins"]
    check("origin panels appear exactly when the plan moves measurable cargo",
          (len(panels) >= 1) == bool(lt["measurable_shipments"]),
          f"{len(panels)} panels for {lt['measurable_shipments']} measured shipments")
    for p in panels:
        label = p["origin"]
        check(f"{label}: unchanged, earlier and later account for every shipment",
              p["unchanged"] + p["earlier"] + p["later"] == p["shipments"],
              f"{p['unchanged']}+{p['earlier']}+{p['later']} vs {p['shipments']}")
        check(f"{label}: the six named outcomes cover the same shipments",
              sum(b["count"] for b in p["buckets"]) == p["shipments"])
        check(f"{label}: the distribution holds every shipment too",
              sum(h["count"] for h in p["histogram"]) == p["shipments"])
        check(f"{label}: the excluded cargo is named rather than counted in",
              (p["cannot_move"] + p["left_alone"] + p["unmeasurable"] + p["shipments"]
               == p["origin_total"]),
              f"{p['cannot_move']}+{p['left_alone']}+{p['unmeasurable']}+{p['shipments']} "
              f"vs {p['origin_total']}")
        check(f"{label}: the panel reports only cargo the plan moves",
              p["shipments"] <= lt["measurable_shipments"])
        check(f"{label}: a worst case is only claimed where something is later",
              (p["worst_days"] > 0) == bool(p["later"]),
              f"worst {p['worst_days']}, later {p['later']}")
    check("no shipment is counted on two panels",
          sum(p["shipments"] for p in panels) <= lt["measurable_shipments"],
          f"{sum(p['shipments'] for p in panels)} vs {lt['measurable_shipments']}")

    risk = result["lead_time_risk"]
    if lt["measurable_shipments"]:
        check("every measured shipment sits on exactly one lane",
              sum(r["shipments"] for r in risk["lanes"]) == lt["measurable_shipments"],
              f"{sum(r['shipments'] for r in risk['lanes'])} vs {lt['measurable_shipments']}")
        check("the lanes' late shipments sum to the run's own count",
              sum(r["late"] for r in risk["lanes"]) == lt["later"] == risk["later"],
              f"{sum(r['late'] for r in risk['lanes'])} vs {lt['later']}")
        check("the worst case named is the worst case measured",
              risk["worst_days"] == lt["later_worst_days"])
        check("a lane counted as clean has no late shipment on it",
              risk["lanes_clean"] == sum(1 for r in risk["lanes"] if not r["late"]))
        check("the concentration claim is measured against volume, not asserted",
              0.0 <= risk["top_lanes_share"] <= 1.0
              and (risk["top_lanes_share"] == 0) == (risk["later"] == 0)
              and risk["concentrated"] == (risk["top_lanes_share"] >= 0.5
                                           and risk["top_lanes_share"]
                                           >= 1.25 * risk["top_lanes_volume_share"]),
              f"top three lanes hold {risk['top_lanes_share']:.0%} of late shipments on "
              f"{risk['top_lanes_volume_share']:.0%} of volume")
        check("each lane's own late count matches its buckets",
              all(sum(b["count"] for b in r["buckets"]) == r["late"]
                  for r in risk["lanes"]))
    else:
        check("an unchanged plan creates no lead-time risk rows",
              not risk["lanes"], f"{len(risk['lanes'])} rows")

    print("\nThe inbound leg is bought the cheaper way, and the arithmetic holds")
    ib = result["inbound"]
    trucks, solo = result["inbound_trucks"], result["inbound_loads_today"]
    used = result["config_used"]
    if ib.get("trucks"):
        check("pooling a region's same-day cargo never needs more loads than not pooling",
              len(trucks) <= len(solo),
              f"{len(trucks)} trailers against {len(solo)} shipment-by-shipment loads")
        over = [t for t in trucks + solo
                if t["pallets"] > used["TRAILER_PALLET_MAX"]
                or t["cbm"] > used["TRAILER_CBM_MAX"] + 1e-6
                or t["gwt"] > used["TRAILER_WEIGHT_MAX_KG"] + 1e-6]
        check("no trailer load breaches the trailer's pallet, volume or payload cap",
              not over, f"{len(over)} breaches")
        check("every trailer's cargo is accounted to the containers it ends up in",
              all(abs(sum(t["containers"].values()) - 1.0) < 1e-9 for t in trucks + solo))

        # The receiving bill is charged per arrival, so the arrivals it is charged on have
        # to be the arrivals the bin-packer found -- not a count taken from anywhere else.
        led = result["ledger"]
        recv = led[led["rate_id"] == "NEW-CFS-INBOUND"]
        arrivals = len(trucks) if ib["path"] == "ftl" else len(solo)
        check("the warehouse is paid to receive exactly the loads that arrive",
              close(float(recv["quantity"].sum()), arrivals, 0.01)
              and close(float(recv["usd"].sum()), arrivals * ib["receiving_rate"], 1.0),
              f"{float(recv['quantity'].sum()):.2f} charged against {arrivals} arrivals, "
              f"${float(recv['usd'].sum()):,.2f}")

        # Collection is the leg the plan does not touch, and this is the check that says
        # so. Charged on the cargo's own inbound load count, it has to come out at what
        # these rates make of the loads the client already runs -- whatever the plan builds.
        if ib["path"] != "ftl":
            alloc = pd.DataFrame(result["allocation"])
            through = set(result["containers_costed"].loc[
                result["containers_costed"]["passthrough"], "container"])
            moved = alloc[~alloc["container"].isin(through) & alloc["term"].eq("EXW")]
            idx = result["shipments"].set_index("grp_key")
            expected = 0.0
            for grp in moved["grp"].unique():
                row = idx.loc[grp]
                rate = result["rate_table"].get(
                    "ORIGIN_COMPONENT", (row["pickup_region"], C.PICKUP_CODE))
                if rate is not None:
                    expected += float(rate["rate_usd"]) * int(row["containers_today"])
            charged = float(led.loc[led["rate_id"].astype(str).str.endswith(
                "-" + C.PICKUP_CODE), "usd"].sum())
            check("collection is charged on the cargo's own loads, not on the plan's boxes",
                  close(charged, expected, max(1.0, 0.002 * max(expected, 1.0))),
                  f"${charged:,.2f} charged against ${expected:,.2f} — what these rates "
                  "make of the loads they already run")

    if spec["ftl"] == "absent":
        # The run's own plan, not the bare board planned earlier: the board is evaluated
        # without volumes to see what the file alone can price, so the ask it carries is
        # zero by construction and checking it would prove nothing.
        alt = result["sourcing"]["origin_collection"].alternative
        check("no trailer rate exists, so the opportunity stays open",
              not ftl_path.exists() and alt["state"] == "opportunity",
              alt.get("state", ""))
        check("the ask is quantified from the shipments they collect today",
              alt["pickups_today"] == int((result["shipments"]["in_scope"]
                                           & result["shipments"]["term"].eq("EXW")).sum()),
              f"{alt.get('pickups_today')}")
        check("and the plan is costed the way they buy the leg today",
              ib["path"] == "groupage" and not ib.get("supplied"))
        check("a haulage target is shown only when the final plan has inbound work",
              ((ib.get("headroom_per_load", 0) > 0) == bool(s["consolidated_lanes"])),
              f"{s['consolidated_lanes']} adopted lanes, "
              f"${ib.get('headroom_per_load', 0):,.2f} a load")
    else:
        with_ftl = pd.concat([rate_card, pd.read_csv(ftl_path)], ignore_index=True)
        ftl_result = run_mod.run(charge_lines, rate_card=with_ftl, site_list=site_list,
                                 answers=answers)
        fib = ftl_result["inbound"]
        check("the tender is tested against both halves of what it would replace",
              fib.get("ftl_total") is not None
              and close(fib["groupage_total"],
                        fib["groupage_haul"] + fib["groupage_receiving"], 1.0)
              and close(fib["ftl_total"], fib["ftl_haul"] + fib["ftl_receiving"], 1.0),
              f"groupage ${fib.get('groupage_total', 0):,.0f} against tender "
              f"${fib.get('ftl_total', 0):,.0f}")
        check("and the cheaper of the two is the one costed",
              (fib["path"] == "ftl") == (fib["ftl_total"] < fib["groupage_total"]),
              f"path {fib['path']}, delta ${fib['delta_total']:,.0f}")
        check("any tender-driven lane change still obeys the commercial rule",
              all(costing_mod.lane_clears_rule(row, ftl_result["config_used"])
                  for _, row in ftl_result["lanes"].loc[
                      ftl_result["lanes"]["verdict"].eq("Consolidate")].iterrows()),
              f"{ftl_result['summary']['containers_future']} boxes with tender against "
              f"{s['containers_future']} without")
        if spec["ftl"] == "adopted":
            check("the cheaper tender is adopted and the saving grows",
                  fib["path"] == "ftl"
                  and ftl_result["summary"]["saving_usd"] > s["saving_usd"],
                  f"${ftl_result['summary']['saving_usd']:,.0f} vs ${s['saving_usd']:,.0f}")
            check("and the plan's inbound money comes off their tender",
                  float(ftl_result["ledger"].loc[
                      ftl_result["ledger"]["rate_id"].eq("NEW-INBOUND-FTL"),
                      "usd"].sum()) > 0)
        else:
            # The point of this scenario. A model that only ever confirms the client's
            # hopes is selling; one that can say "we tested it and yours is cheaper" is
            # advising, and the saving must not move by a cent.
            check("the dearer tender is rejected", fib["path"] == "groupage",
                  f"${fib['delta_total']:,.0f} dearer across the year")
            check("rejecting it leaves the saving untouched",
                  close(ftl_result["summary"]["saving_usd"], s["saving_usd"], 0.01),
                  f"${ftl_result['summary']['saving_usd']:,.2f} vs ${s['saving_usd']:,.2f}")
            check("and nothing is charged on a rate the model rejected",
                  not len(ftl_result["ledger"][
                      ftl_result["ledger"]["rate_id"].eq("NEW-INBOUND-FTL")]))
        check("every control still passes with the tender in play",
              ftl_result["controls_summary"]["all_passed"],
              ", ".join(ftl_result["controls_summary"]["failures"]))

    print("\nSecond run: hand the mappings back and the review disappears")
    reused = run_mod.run(charge_lines, rate_card=rate_card,
                         site_mapping=result["resolution"].sites,
                         supplier_mapping=result["resolution"].suppliers)
    check("the queue is empty when mappings are supplied",
          len(reused["queue"]) == 0, f"{len(reused['queue'])} items")
    check("the answer is identical to the first run",
          close(reused["summary"]["saving_usd"], s["saving_usd"], 0.01)
          and reused["summary"]["containers_future"] == s["containers_future"],
          f"${reused['summary']['saving_usd']:,.2f} vs ${s['saving_usd']:,.2f}")

    print("\nThe consolidation service must be priced by the client, or there is no run")
    #
    # The single most important refusal in the model. The service is cost that exists
    # only after consolidating, so pricing it from a benchmark of ours put most of the
    # saving on figures the client had never seen, and leaving it out would be worse
    # still -- the saving would grow the less they had told us. Neither is allowed.
    try:
        run_mod.run(charge_lines, site_list=site_list, answers=answers)
        check("no consolidation quote, no answer", False, "it ran anyway")
    except run_mod.MissingServiceQuote as exc:
        check("no consolidation quote, no answer", True, "")
        check("and the refusal names the legs it cannot price",
              len(exc.labels) >= 3 and all(isinstance(x, str) for x in exc.labels),
              ", ".join(exc.labels))
    check("nothing in the priced model rests on a benchmark of ours",
          s["provenance"]["rates_assumed"] == 0
          and s["provenance"]["usd_assumed"] == 0
          and "benchmark" not in set(result["service_pricing"].values()),
          f"{s['provenance']['rates_assumed']} rates, "
          f"${s['provenance']['usd_assumed']:,.0f}")

    print("\nAssumptions move the answer, and a client's own quote outranks them")
    # Two halves of one rule, and both directions matter.
    #
    # With a consolidation quote uploaded, the quote is the number -- so editing the
    # settings field must change nothing, or the interface and the model would disagree
    # about the same figure with nothing on screen saying which one won. Without a quote,
    # the field is the only source there is and must move the answer, or the control is
    # inert and the whole Settings step is decoration.
    fixed = run_mod.run(charge_lines, rate_card=rate_card, site_list=site_list,
                        answers=answers,
                        overrides={"CFS_HANDLING_PER_CONTAINER": 2000.0})
    check("an uploaded quote outranks the settings field",
          close(fixed["summary"]["saving_usd"], s["saving_usd"], 0.01),
          f"${fixed['summary']['saving_usd']:,.2f} vs ${s['saving_usd']:,.2f} — the "
          "client's quoted handling rate must not be editable behind their back")

    # ...and the other half: a figure the client types is theirs, so it outranks even the
    # quote they uploaded. Without this the settings step is decoration, and worse, the
    # screen would show one number while the model costed another.
    edited = run_mod.run(
        charge_lines, rate_card=rate_card, site_list=site_list, answers=answers,
        overrides={"CFS_HANDLING_PER_CONTAINER": 2000.0},
        service_pricing={"CFS_HANDLING_PER_CONTAINER": "edited"})
    handling = edited["rate_table"].get("SERVICE", "cfs_handling")
    check("a figure they typed outranks the quote in the rate table",
          handling["rate_usd"] == 2000.0
          and (edited["summary"]["saving_usd"] < s["saving_usd"]
               or not s["consolidated_lanes"]),
          f"costed ${handling['rate_usd']:,.0f}, saving "
          f"${edited['summary']['saving_usd']:,.0f} vs ${s['saving_usd']:,.0f}")
    check("and it is recorded as theirs, not as a benchmark of ours",
          handling["source"] == "QUOTED_NOT_YET_BOUGHT"
          and "You entered this figure" in handling["derivation"],
          handling["source"])

    # No decision may offer the same answer twice. "Add as a new site: Kent, UK" and
    # "Treat as a single site: Kent, UK" both resolved to a site called Kent, UK, so the
    # reviewer was choosing between two spellings of one answer -- work with no decision
    # attached to it.
    dupes = []
    for item in resolution.queue:
        seen = {}
        for opt in item.options:
            target = opt
            for prefix in ("Add as a new site: ", "Merge into "):
                if target.startswith(prefix):
                    target = target[len(prefix):]
            seen.setdefault(target, []).append(opt)
        dupes += [(item.key, v) for v in seen.values() if len(v) > 1]
    check("no decision offers two options that mean the same thing",
          not dupes, "; ".join(f"{k}: {v}" for k, v in dupes[:2]))

    # The lane table has to add up to the run it describes, under *either* pool key.
    # It did not: lanes were always cut by delivery site while the packer pools by
    # country when sites may mix, so every lane found zero modelled containers and
    # reported its entire cost as saved -- "89 of 89 containers removed, boxes filling to
    # 0 CBM" on every row, with a correct headline above it.
    for pool_key in ("cfs_pod_site", "cfs_pod"):
        pooled = run_mod.run(charge_lines, rate_card=rate_card, site_list=site_list,
                             answers=answers, overrides={"POOL_KEY": pool_key})
        L, ps = pooled["lanes"], pooled["summary"]
        check(f"[{pool_key}] every runtime control passes",
              pooled["controls_summary"]["all_passed"],
              ", ".join(pooled["controls_summary"]["failures"]))
        check(f"[{pool_key}] the lanes account for every container in the plan",
              close(float(L["containers_future"].sum()), ps["containers_future"], 0.05),
              f"{L['containers_future'].sum():.2f} across lanes vs "
              f"{ps['containers_future']} built")
        check(f"[{pool_key}] the lanes account for the whole saving",
              close(float(L["saving_usd"].sum()), ps["saving_usd"], 1.0),
              f"${L['saving_usd'].sum():,.0f} across lanes vs ${ps['saving_usd']:,.0f}")
        check(f"[{pool_key}] no lane carries cargo with no container behind it",
              not len(L[(L["containers_future"] == 0) & (L["pallets"] > 0)]),
              f"{len(L[(L['containers_future'] == 0) & (L['pallets'] > 0)])} such lanes")

    # A lane whose box count goes UP is either LCL cargo moving into the client's own
    # containers -- which is real and has to be explained on screen -- or a packing bug.
    # There is no third possibility, so it is asserted rather than eyeballed.
    lanes = result["lanes"]
    risen = lanes[lanes["containers_saved"] < 0]
    check("a lane only uses more containers where LCL cargo moves into them",
          bool((risen["lcl_cbm_today"] > 0).all()),
          f"{len(risen)} lanes rise, "
          f"{int((risen['lcl_cbm_today'] <= 0).sum())} of them with no LCL behind it")
    check("and each of those lanes says so in plain words",
          all("ships LCL today" in w for w in risen["why"]),
          f"{len(risen)} lanes checked")
    check("the volume adds up: every CBM is either in a box today or shipped LCL",
          close(float(result["shipments"].loc[result["shipments"]["in_scope"], "cbm"].sum()),
                float(lanes["cbm"].sum()), 0.5),
          f"{float(lanes['cbm'].sum()):,.1f} CBM across lanes")

    tighter = run_mod.run(charge_lines, rate_card=rate_card, site_list=site_list,
                          answers=answers, overrides={"MAX_DWELL_DAYS": 3})
    check("a tighter dwell cap builds more containers",
          tighter["summary"]["containers_future"] >= s["containers_future"],
          f"{tighter['summary']['containers_future']} vs {s['containers_future']}")
    check("the tighter-dwell route still reaches results with every control passing",
          tighter["controls_summary"]["all_passed"],
          ", ".join(tighter["controls_summary"]["failures"]))
    # The interface describes the download off workbook.SHEETS. If the writer stops
    # matching that list, the description on the results screen becomes a claim about a
    # file the client is holding and can count the tabs of.
    import io

    import workbook as workbook_mod
    from openpyxl import load_workbook
    buf = io.BytesIO()
    workbook_mod.build_workbook(buf, result)
    buf.seek(0)
    written = load_workbook(buf).sheetnames
    check("the workbook writes exactly the tabs the app promises",
          written == [name for name, _ in workbook_mod.SHEETS],
          f"{written} vs {[n for n, _ in workbook_mod.SHEETS]}")

    cfg_check = C.values({"OUT_CBM_MAX": 60.0, "OUT_CBM_TARGET_PCT": 0.80})
    check("a dispatch target can never exceed its cap",
          cfg_check["OUT_CBM_TARGET"] <= cfg_check["OUT_CBM_MAX"]
          and abs(cfg_check["OUT_CBM_TARGET"] - 48.0) < 0.01,
          f"target {cfg_check['OUT_CBM_TARGET']} against cap {cfg_check['OUT_CBM_MAX']}")

    headline = {
        "client": manifest["client"],
        "containers": f"{s['containers_today']} -> {s['containers_future']}",
        "reduction": f"{s['container_reduction_pct']:.1%}",
        "saving": f"${s['saving_usd']:,.0f}  ({s['saving_pct']:.1%})",
        "fill": f"{s['today_fill_cbm']} -> {s['mean_fill_cbm']} CBM",
        "rates": (f"{s['provenance']['rates_from_card']} card, "
                  f"{s['provenance']['rates_derived']} derived, "
                  f"{s['provenance']['rates_quoted']} quoted, "
                  f"{s['provenance']['rates_assumed']} benchmark, 0 invented"),
        "evidenced": f"{s['provenance']['evidenced_share']:.1%}",
        "delivery": (f"+{lt['mean_delta_days_conservative']} days mean, "
                     f"{lt['unaffected_pct']:.0%} unaffected"),
        "queue": f"{len(resolution.queue)} decisions, "
                 f"{resolution.stats['auto_resolved']} auto-resolved",
        "board (no cards)": ", ".join(f"{v} {k}" for k, v in
                                      sorted(bare_board["by_state"].items())),
        "ftl": spec["ftl"],
    }
    return check, headline


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--scenario", choices=sorted(SCENARIOS))
    args = ap.parse_args()

    wanted = [args.scenario] if args.scenario else list(SCENARIOS)
    results = []
    for scenario_id in wanted:
        if not (DATA / scenario_id / "ground_truth_manifest.json").exists():
            print(f"\n{scenario_id}: no data — run generator/make_demo_data.py first")
            return 1
        results.append(validate_scenario(scenario_id, SCENARIOS[scenario_id]))

    print("\n" + "=" * 74)
    print("  What each demo will say")
    print("=" * 74)
    for check, headline in results:
        print(f"\n  {headline.pop('client')}")
        for key, value in headline.items():
            print(f"    {key:<18} {value}")
    print("\n" + "-" * 74)

    total = sum(len(c.rows) for c, _ in results)
    failed = [(c.label, n) for c, _ in results for n in c.failed]
    print(f"\n{total - len(failed)} of {total} checks passed "
          f"across {len(results)} dataset{'s' if len(results) != 1 else ''}")
    if failed:
        print("\nFAILED:")
        for label, name in failed:
            print(f"  - [{label}] {name}")
        return 1
    print("Every demo is safe to run.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
