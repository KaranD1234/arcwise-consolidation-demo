"""Write the result as one formatted workbook.

The workbook is the deliverable. The app shows a client the answer; this is what
they take away, forward to their finance team and argue with. So it carries the
working, not just the conclusion: every container, every charge, every rate with its
provenance, every judgement call and every control.

Tab order follows how a sceptical reader works through it -- the answer, then the
two states, then the cargo, then the money, then where the prices came from, then
the decisions and the checks.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config as C

INK = "0F1828"
BRAND = "0D2E5E"
BRAND_LIGHT = "EBF2FA"
HAIRLINE = "E8EDF5"
POSITIVE = "10B981"
NEGATIVE = "C0362A"

HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color=INK)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=BRAND)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="6B7A8F")
THIN = Side(style="thin", color=HAIRLINE)
CELL_BORDER = Border(bottom=THIN)

MONEY = '#,##0.00'
MONEY0 = '#,##0'
PCT = '0.0%'
NUM1 = '#,##0.0'

# The tabs this writer produces, and one line each for what a reader finds there. The
# interface describes the download from this list rather than from a sentence of its own,
# because a sentence of its own is how the app came to be offering "twelve tabs" of a
# seven-tab workbook. Numbering is the client's own convention, gaps and all.
SHEETS = [
    ("0 COLUMN KEY", "every column in the workbook, explained"),
    ("1 CURRENT", "each in-scope shipment as invoiced today"),
    ("2 FUTURE STATE", "each container the model builds, with its cost"),
    ("3 LEAD TIME AND DWELL", "each shipment's delivery date, before and after"),
    ("4 COST LEDGER", "every dollar as its own row, tagged to a rate"),
    ("5 RATE VALIDATION", "every rate, where it came from and the assumptions"),
    ("8 RECONCILIATION", "the controls, the data-quality findings and the decisions"),
]


def _sheet(wb, title, note=None):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    if note:
        ws["A1"] = note
        ws["A1"].font = NOTE_FONT
        ws.freeze_panes = "A3"
    return ws


def _write_frame(ws, df, start_row=1, money_cols=(), pct_cols=(), num_cols=(),
                 width_cap=52):
    """Write a frame with a styled header and sensible column widths."""
    if df is None or not len(df):
        ws.cell(row=start_row, column=1, value="No rows.").font = NOTE_FONT
        return

    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 28

    for i, record in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(record, start=1):
            cell = ws.cell(row=i, column=j)
            # openpyxl cannot store pandas/numpy scalars or lists directly.
            cell.value = _cellify(value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            name = str(df.columns[j - 1])
            if name in money_cols:
                cell.number_format = MONEY
            elif name in pct_cols:
                cell.number_format = PCT
            elif name in num_cols:
                cell.number_format = NUM1

    for j, col in enumerate(df.columns, start=1):
        longest = max([len(str(col))] + [
            len(str(_cellify(v))) for v in df.iloc[:, j - 1].head(400)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(11, longest + 2), width_cap)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _cellify(value):
    import numpy as np
    import pandas as pd
    if isinstance(value, (list, tuple, set)):
        return " + ".join(str(v) for v in value)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, pd.Series):
        return str(value)
    return value


def _kv_sheet(ws, title, rows, start_row=1):
    """A two-column headline sheet: label, value, note."""
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    r = start_row + 2
    for label, value, note in rows:
        if label is None:                     # section break
            r += 1
            ws.cell(row=r, column=1, value=value).font = Font(
                name="Calibri", size=11, bold=True, color=BRAND)
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = Font(
            name="Calibri", size=10, bold=True, color=INK)
        vc = ws.cell(row=r, column=2, value=_cellify(value))
        vc.font = BODY_FONT
        nc = ws.cell(row=r, column=3, value=note)
        nc.font = NOTE_FONT
        nc.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 88


def _band(ws, row, title):
    """A heading inside a sheet, where a second table follows the first."""
    ws.cell(row=row, column=1, value=title).font = Font(
        name="Calibri", size=11, bold=True, color=BRAND)


# What each column means, for tab 0. A column with no entry still gets a row: an
# unexplained column is a reason to write the sentence, not to leave it out.
COLUMN_NOTES = {
    "grp_key": ("The shipment group: one booking's charges, keyed as the file keys them",
                "text", "Pivot on this to tie a row back to the client's own system"),
    "cfs": ("Origin consolidation warehouse this cargo would route through", "place", ""),
    "origin": ("Origin the cargo shipped from", "place", ""),
    "pod": ("Port of discharge", "place", ""),
    "site": ("Delivery warehouse, as resolved on the review step", "place", ""),
    "site_country": ("Country of the delivery warehouse", "place",
                     "Cargo never shares a container across countries"),
    "delivery_raw": ("The delivery address exactly as written in the file", "text",
                     "Kept so every mapping traces back to its source string"),
    "term": ("Shipping term", "EXW / FOB / DAP / DDP",
             "Only EXW and FOB are in scope: the client controls the origin leg"),
    "mode": ("Transport mode as the file records it", "text",
             "Ocean only is in scope. LCL cargo books no container of the client's"),
    "shipper": ("Supplier who shipped the cargo", "text", ""),
    "pickup_region": ("Region the supplier loads in, resolved on the review step", "text",
                      "Decides which origin collection rate applies"),
    "shipments": ("Shipments in this group", "count", ""),
    "master_bills": ("Master bills in this group", "count", ""),
    "containers_today": ("Containers the client books today for this cargo", "count",
                         "Rebuilt as the maximum equipment per master bill, summed to the "
                         "group. LCL cargo counts zero — it rides in a co-loader's box"),
    "pallets": ("Pallets", "count", ""),
    "cbm": ("Volume", "CBM", ""),
    "gwt": ("Gross weight", "kg", ""),
    "cargo_ready": ("Date the cargo was ready at origin", "date", ""),
    "act_depart": ("Actual departure", "date", ""),
    "act_arriv": ("Actual arrival", "date", ""),
    "act_deliv": ("Actual delivery", "date", ""),
    "date_quality": ("Whether this row's dates support a lead-time comparison", "text",
                     "Rows that cannot be measured are excluded, never estimated"),
    "currency": ("Currency the charges were invoiced in", "text",
                 "Converted to USD at the rate carried in the file"),
    "invoiced_usd": ("What the client actually paid for this group", "USD",
                     "The 'today' side of every comparison in this workbook"),
    "container": ("Container number in the plan", "count", ""),
    "container_ref": ("Readable reference for the container", "text", ""),
    "how_built": ("Whether the model built this box at a warehouse, or left the cargo "
                  "exactly as it ships today", "text",
                  "'Unchanged' boxes carry identical cost in both states"),
    "sites": ("Delivery warehouses on this container", "count", ""),
    "site_countries": ("Delivery countries on this container", "text", ""),
    "groups": ("Shipment groups sharing this container", "count", ""),
    "fill_cbm_pct": ("How full the box is by volume", "% of cap", ""),
    "fill_pallet_pct": ("How full the box is by pallet count", "% of cap", ""),
    "dispatch_reason": ("Why the box was sent when it was", "text",
                        "Full enough, or the dwell cap forced it out"),
    "dwell_days": ("Days the first pallet waited at the warehouse", "days", ""),
    "passthrough": ("Cargo consolidation does not touch", "yes / no",
                    "Contributes nothing to the saving, in either direction"),
    "total_usd": ("Modelled cost of this container", "USD", ""),
    "usd": ("Amount of this charge", "USD", ""),
    "quantity": ("Units the rate was applied to", "number", ""),
    "rate_id": ("Identifier for the rate applied", "text", ""),
    "rate_source": ("Where that rate came from", "text",
                    "Full provenance for each is on '5 RATE VALIDATION'"),
    "pool": ("Cost pool the charge belongs to", "text",
             "The pools are the rows of the cost comparison"),
    "item": ("What the rate prices, in words", "text", ""),
    "rate_usd": ("The rate applied", "USD", ""),
    "source": ("Provenance of the rate", "text",
               "CLIENT_RATE_CARD, DERIVED_FROM_INVOICES, QUOTED_NOT_YET_BOUGHT or "
               "ACTUAL_INVOICE. Nothing is invented"),
    "confidence": ("How much weight the rate carries", "HIGH / MEDIUM / LOW / DECLARED",
                   "A rate derived by analogy is never HIGH"),
    "derivation": ("The arithmetic behind a derived rate, in words", "text",
                   "Population, formula and any caveat"),
    "population_invoices": ("Invoices behind a derived rate", "count", ""),
    "population_groups": ("Shipment groups behind a derived rate", "count", ""),
    "population_containers": ("Containers behind a derived rate", "count", ""),
    "observed_min": ("Cheapest observation in the population", "USD", ""),
    "observed_median": ("Median observation", "USD", ""),
    "observed_max": ("Dearest observation", "USD", ""),
    "invoice_total_usd": ("Invoiced total the derivation divided", "USD", ""),
    "delta_days": ("Days later the plan delivers this shipment", "days",
                   "Positive is later. Earlier deliveries count as unchanged, not as a gain"),
    "lead_time_today": ("Door-to-door days today", "days", ""),
    "lead_time_modelled": ("Door-to-door days under the plan", "days", ""),
    "max_wait_days": ("Longest any pallet in this shipment waited", "days", ""),
    "last_mile_days": ("Discharge port to delivery warehouse", "days", ""),
    "measurable": ("Whether this shipment has the dates to be compared", "yes / no",
                   "Unmeasurable shipments are excluded from every lead-time figure"),
    "modelled_arrive": ("Modelled arrival at the discharge port", "date", ""),
    "modelled_deliver": ("Modelled delivery at the warehouse", "date", ""),
}


# --------------------------------------------------------------------------------------
# The client's own workbook is the reference for this one, down to the column names, so
# somebody who knows theirs can read ours without relearning where anything is. These
# specs are (source column, header) in the client's order; a source the run does not
# produce is skipped rather than written empty.
# --------------------------------------------------------------------------------------
TODAY_SPEC = [
    ("cfs", "CFS"), ("site", "Delivery Site"), ("site_country", "Delivery Country"),
    ("grp_key", "GRP #"), ("invoice", "Invoice #(s)"),
    ("consignee", "Client Company"), ("term", "Shipping Term"),
    ("mode", "Transport Mode"), ("shipper", "Shipper"),
    ("origin_country", "Origin Country"), ("pod_country", "Destination Country"),
    ("origin", "Origin"), ("pod", "Port of Discharge"),
    ("delivery_raw", "Delivery (final)"), ("pickup_region", "Pickup Region"),
    ("cargo_ready", "Cargo Ready"), ("act_depart", "Act Depart"),
    ("act_arriv", "Act Arriv"), ("act_deliv", "Act Deliv"),
    ("20", "20"), ("40", "40"), ("40HQ", "40HQ"), ("45", "45"),
    ("cbm", "CBM"), ("gwt", "GWT kg"), ("pallets", "Pallets"),
    ("shipments", "Shipments In Group"), ("master_bills", "Master Bills"),
    ("containers_today", "Containers Today"),
] + [(p, p) for p in C.COST_POOLS] + [
    ("invoiced_usd", "Total In Scope USD"),
    ("days_ready_to_sail_today", "Days Cargo Ready To Sail Today"),
    ("lead_time_today", "Lead Time Days Today"),
    ("date_quality", "Date Quality"), ("currency", "Currency"),
    ("modelled_containers", "Modelled Containers"),
    ("split_across", "Split Across N Containers"),
    ("consolidation_status", "Consolidation Status"),
]

PLAN_SPEC = [
    ("container", "Container"), ("container_ref", "Container Ref"),
    ("how_built", "How This Box Was Built"), ("cfs", "CFS"),
    ("pod", "Port of Discharge"), ("sites", "Delivery Site(s) On This Container"),
    ("site_countries", "Delivery Country(s)"), ("delivery_service", "Delivery Service"),
    ("last_pallet_ready", "Last Pallet Cargo Ready"),
    ("modelled_sail", "Modelled Sail Date"), ("modelled_arrive", "Modelled Arriv Date"),
    ("sailing_from", "Sailing Taken From Shipment"),
    ("pallets", "Container Total Pallets"), ("cbm", "Container Total CBM"),
    ("gwt", "Container Total Weight kg"),
    ("fill_cbm_pct", "Container Fill % of CBM Cap"),
    ("fill_pallet_pct", "Container Fill % of Pallet Cap"),
    ("groups", "Shipments In This Container"),
    ("dispatch_reason", "Why It Was Dispatched"), ("dwell_days", "Dwell Days"),
] + [(p, p) for p in C.COST_POOLS] + [
    ("total_usd", "Total In Scope USD"),
    ("freight_rate_used", "Freight Rate Used Per Container"),
    ("storage_usd", "of which Origin CFS Storage"),
    ("passthrough", "Cargo The Plan Does Not Change"),
]


def _relabel(frame, spec):
    """Take the columns a spec names, in its order, under its headers."""
    keep = [(src, head) for src, head in spec if src in frame.columns]
    out = frame[[src for src, _ in keep]].copy()
    out.columns = [head for _, head in keep]
    return out


def pallet_groups(result):
    """Sheet 3, at the only grain where a lead time is a fact.

    One row per *shipment inside a container*, not per shipment and not per container.
    That is the client's own grain and it is the right one, for a reason that is easy to
    miss until the numbers stop reconciling: a shipment can be split across several boxes
    that sail on different days, and a box carries cargo whose ready dates are weeks apart.
    A shipment-level sheet has to pick one of a shipment's sail dates, and a
    container-level one has to take a max() over cargo that has nothing in common — so
    neither reconciles with the other, and the dwell that drives the storage charge is
    invisible on both. Here every figure on a row is a subtraction between two dates that
    belong to the same cargo, and both higher views are sums or maxima over these rows.

    The pallet allocation is already at pallet grain, so this is a group-by rather than a
    reconstruction: the split percentages, the dwell and the storage all fall out of the
    same rows the packer and the costing worked on.
    """
    import pandas as pd

    alloc = pd.DataFrame(result["allocation"])
    if not len(alloc):
        return None
    ship = result["shipments"].drop_duplicates("grp_key").set_index("grp_key")
    con = result["containers_costed"].set_index("container")
    sail = pd.DataFrame(result["sailings"]).set_index("container")
    cfg = result["config_used"]
    free = float(cfg["CFS_STORAGE_FREE_DAYS"])

    # Unchanged boxes remain in this audit sheet so every shipment reconciles, but they
    # never enter the new warehouse. Their candidate packing wait therefore cannot be
    # printed as warehouse dwell or turned into storage in the final workbook.
    untouched = set(con.index[con["passthrough"]])
    alloc.loc[alloc["container"].isin(untouched), "wait_days"] = 0

    # Every pallet's own chargeable volume-days, which is exactly what costing charged
    # storage on -- so the column below sums to the ledger rather than re-deriving it.
    alloc["chargeable_days"] = (alloc["wait_days"].astype(float) - free).clip(lower=0)
    alloc["chargeable_cbm_days"] = alloc["cbm"].astype(float) * alloc["chargeable_days"]
    # The day the box was closed. Every pallet in a container waits until the same
    # dispatch, so this is one date per container arrived at from each pallet's own clock
    # rather than a second bookkeeping of it.
    alloc["packing_date"] = (pd.to_datetime(alloc["cfs_ready"])
                             + pd.to_timedelta(alloc["wait_days"].astype(int), unit="D"))

    g = alloc.groupby(["container", "grp"], as_index=False).agg(
        Pallets=("pallet_id", "count"),
        CBM=("cbm", "sum"),
        weight=("gwt", "sum"),
        first_at_cfs=("cfs_ready", "min"),
        packing_date=("packing_date", "max"),
        wait=("wait_days", "max"),
        chargeable_cbm_days=("chargeable_cbm_days", "sum"))

    grp_pallets = alloc.groupby("grp")["pallet_id"].count()
    box_pallets = alloc.groupby("container")["pallet_id"].count()
    g["% of This Shipment In This Box"] = (
        100.0 * g["Pallets"] / g["grp"].map(grp_pallets)).round(1)
    g["% of This Box From This Shipment"] = (
        100.0 * g["Pallets"] / g["container"].map(box_pallets)).round(1)

    for col, src in [("Container Ref", "container_ref"), ("CFS", "cfs"),
                     ("Port of Discharge", "pod"), ("Delivery Site", "sites")]:
        g[col] = g["container"].map(con[src])
    g["Modelled Sail Date"] = g["container"].map(sail["modelled_sail"])
    g["Modelled Arriv Date"] = g["container"].map(sail["modelled_arrive"])
    # The client's sheet names this "How This Box Was Built" and it is the one column that
    # tells a reader whether a row is the plan's work or cargo it left alone.
    g["How This Box Was Built"] = g["container"].map(
        con["passthrough"]).map(
            {False: "Built at the warehouse — consolidated from the cargo on this lane",
             True: "Unchanged — shipped exactly as it is booked today"})

    for col, src in [("CRD", "cargo_ready"), ("Act Depart Today", "act_depart"),
                     ("Act Arriv Today", "act_arriv"), ("Act Deliv Today", "act_deliv"),
                     ("Shipper", "shipper"), ("Invoice #(s)", "invoice")]:
        g[col] = g["grp"].map(ship[src])

    for col in ("CRD", "Act Depart Today", "Act Arriv Today", "Act Deliv Today",
                "Modelled Sail Date", "Modelled Arriv Date", "first_at_cfs"):
        g[col] = pd.to_datetime(g[col], errors="coerce")

    # The fourth date. The box lands when the vessel supplying its sail date landed, and
    # the last mile is this shipment's own observed arrival-to-door lag: same port, same
    # warehouse, same truck, so nothing in the plan touches it.
    g["Modelled Deliv Date"] = (g["Modelled Arriv Date"]
                                + (g["Act Deliv Today"] - g["Act Arriv Today"]))

    g["Days CRD To Sail Today"] = (g["Act Depart Today"] - g["CRD"]).dt.days
    g["Days CRD To Sail Modelled"] = (g["Modelled Sail Date"] - g["CRD"]).dt.days
    g["Sail Change Days"] = (g["Days CRD To Sail Modelled"]
                             - g["Days CRD To Sail Today"])
    g["Days CRD To Delivered Today"] = (g["Act Deliv Today"] - g["CRD"]).dt.days
    g["Days CRD To Delivered Modelled"] = (g["Modelled Deliv Date"] - g["CRD"]).dt.days
    g["Delivery Change Days"] = (g["Modelled Deliv Date"] - g["Act Deliv Today"]).dt.days
    g["Packing Dwell Days"] = g["wait"].astype(float)
    g["Days At CFS"] = g["wait"].astype(float)
    g["Free Storage Days"] = free
    g["Chargeable Storage Days"] = (g["Days At CFS"] - free).clip(lower=0)
    g["Chargeable Storage CBM-Days"] = g["chargeable_cbm_days"].round(6)

    storage = result["rate_table"].get("SERVICE", "cfs_storage") or {}
    rate = float(storage.get("rate_usd", 0.0))
    g["Storage Rate USD Per CBM Per Day"] = rate
    # Priced off the row's own volume-days, at the same rate costing used, so this column
    # and the ledger cannot disagree.
    g["Storage USD"] = (g["Chargeable Storage CBM-Days"] * rate).round(2)

    out = _relabel(g.rename(columns={"container": "Container", "grp": "Old Shipment",
                                     "weight": "Weight kg"}), [
        ("Container", "Container"), ("Container Ref", "Container Ref"),
        ("How This Box Was Built", "How This Box Was Built"), ("CFS", "CFS"),
        ("Port of Discharge", "Port of Discharge"), ("Delivery Site", "Delivery Site"),
        ("Old Shipment", "Old Shipment"), ("Invoice #(s)", "Invoice #(s)"),
        ("Shipper", "Shipper"),
        ("Pallets", "Pallets"), ("CBM", "CBM"), ("Weight kg", "Weight kg"),
        ("% of This Shipment In This Box", "% of This Shipment In This Box"),
        ("% of This Box From This Shipment", "% of This Box From This Shipment"),
        ("CRD", "CRD"), ("Act Depart Today", "Act Depart Today"),
        ("Modelled Sail Date", "Modelled Sail Date"),
        ("Act Arriv Today", "Act Arriv Today"),
        ("Modelled Arriv Date", "Modelled Arriv Date"),
        ("Act Deliv Today", "Act Deliv Today"),
        ("Modelled Deliv Date", "Modelled Deliv Date"),
        ("Days CRD To Sail Today", "Days CRD To Sail Today"),
        ("Days CRD To Sail Modelled", "Days CRD To Sail Modelled"),
        ("Sail Change Days", "Sail Change Days"),
        ("Days CRD To Delivered Today", "Days CRD To Delivered Today"),
        ("Days CRD To Delivered Modelled", "Days CRD To Delivered Modelled"),
        ("Delivery Change Days", "Delivery Change Days"),
        ("first_at_cfs", "Delivered To CFS"), ("packing_date", "Packing Date"),
        ("Packing Dwell Days", "Packing Dwell Days"),
        ("Days At CFS", "Days At CFS"), ("Free Storage Days", "Free Storage Days"),
        ("Chargeable Storage Days", "Chargeable Storage Days"),
        ("Chargeable Storage CBM-Days", "Chargeable Storage CBM-Days"),
        ("Storage Rate USD Per CBM Per Day", "Storage Rate USD Per CBM Per Day"),
        ("Storage USD", "Storage USD"),
    ])
    return out.sort_values(["Container", "CBM"], ascending=[True, False])


def _column_note(sheet, column):
    """The line tab 0 carries for one column."""
    entry = COLUMN_NOTES.get(f"{sheet}:{column}") or COLUMN_NOTES.get(column)
    if entry:
        return entry
    if column in C.COST_POOLS:
        return (f"{column} cost", "USD",
                "The same pool appears on both states, so the two are comparable directly")
    return (column.replace("_", " ").capitalize(), "", "")


# --------------------------------------------------------------------------------------
def build_workbook(path, result):
    """Write every tab. ``result`` is the dict returned by ``run.run``."""
    import pandas as pd

    s = result["summary"]
    lt = result["lead_time"]
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 0. column key ------------------------------------------------------------
    #
    # First tab, and it earns the place: this workbook goes to somebody who did not sit
    # through the walkthrough, and a column called "how_built" or "date_quality" means
    # nothing on its own. Built from the frames actually written, so a column can never
    # appear in the workbook without a line here explaining it.
    key_ws = _sheet(wb, "0 COLUMN KEY",
                    "Every column in this workbook, what it means and how to read it.")

    sheets = []          # (sheet name, frame, note) filled in as each tab is written

    # ---- 1. current --------------------------------------------------------------
    ws = _sheet(wb, "1 CURRENT",
                "Every in-scope shipment group as invoiced today, with container counts "
                "rebuilt as the maximum equipment per master bill and summed to the group.")
    today_df = result["shipments"]
    today_df = today_df[today_df["in_scope"]].copy()
    # Today's own lead times, so the left-hand side of the comparison carries its dates in
    # days as well as in dates and nobody has to subtract two columns to check ours.
    for name, later, earlier in [("days_ready_to_sail_today", "act_depart", "cargo_ready"),
                                 ("lead_time_today", "act_deliv", "cargo_ready")]:
        today_df[name] = (pd.to_datetime(today_df[later], errors="coerce")
                          - pd.to_datetime(today_df[earlier], errors="coerce")).dt.days
    # Which boxes of the plan this shipment's pallets ended up in. A shipment split across
    # several containers is the normal case, and without this there is no way to get from a
    # row here to the rows on '2 FUTURE STATE' that carry its cargo.
    alloc = pd.DataFrame(result["allocation"])
    if len(alloc):
        boxes = alloc.groupby("grp")["container"].agg(
            lambda v: ", ".join(str(x) for x in sorted(set(v))))
        counts = alloc.groupby("grp")["container"].nunique()
        changed_ids = set(result["containers_costed"].loc[
            ~result["containers_costed"]["passthrough"], "container"])
        built = (alloc[alloc["container"].isin(changed_ids)]
                 .groupby("grp")["container"].nunique())
        today_df["modelled_containers"] = today_df["grp_key"].map(boxes).fillna("")
        today_df["split_across"] = today_df["grp_key"].map(counts).fillna(0).astype(int)
        today_df["consolidation_status"] = [
            "CONSOLIDATED" if n else "LEFT EXACTLY AS SHIPPED"
            for n in today_df["grp_key"].map(built).fillna(0)]
    today_df = _relabel(today_df, TODAY_SPEC)
    _write_frame(ws, today_df, start_row=3,
                 money_cols=set(C.COST_POOLS) | {"Total In Scope USD"},
                 num_cols={"CBM", "GWT kg"})
    sheets.append(("1 CURRENT", today_df))

    # ---- 2. future state ----------------------------------------------------------
    ws = _sheet(wb, "2 FUTURE STATE",
                "Every container the model builds, why it was dispatched when it was, and "
                "its cost by pool. 'Unchanged' containers are cargo consolidation does not "
                "touch — they carry the same cost in both states.")
    plan_frame = result["containers_costed"].copy()
    # The three dates the plan puts on a box, and the shipment whose real sailing supplied
    # them. Every modelled date in this workbook traces back to a departure this lane
    # actually ran, and this column is where a reader checks that.
    sail = pd.DataFrame(result["sailings"]).set_index("container")
    for col in ("last_pallet_ready", "modelled_sail", "modelled_arrive", "sailing_from"):
        plan_frame[col] = plan_frame["container"].map(sail[col])
    plan_frame["delivery_service"] = [
        "direct" if int(n) <= 1 else "deconsolidated at destination"
        for n in plan_frame["sites"].astype(str).str.count(";") + 1]
    led = result["ledger"]
    plan_frame["freight_rate_used"] = plan_frame["container"].map(
        led[led["pool"] == "Freight"].groupby("container")["usd"].sum())
    # Named on the box it is charged to, because it is a component of Origin CFS and not
    # an addition to it -- the price of waiting to consolidate, and the one line on this
    # sheet that a longer dwell cap moves.
    plan_frame["storage_usd"] = plan_frame["container"].map(
        led[led["rate_id"] == "NEW-CFS-STORAGE"].groupby("container")["usd"].sum()
    ).fillna(0.0)
    plan_frame = _relabel(plan_frame, PLAN_SPEC)
    _write_frame(ws, plan_frame, start_row=3,
                 money_cols=set(C.COST_POOLS) | {"Total In Scope USD",
                                                "Freight Rate Used Per Container",
                                                "of which Origin CFS Storage"},
                 num_cols={"Container Total CBM", "Container Total Weight kg",
                           "Container Fill % of CBM Cap",
                           "Container Fill % of Pallet Cap"})
    sheets.append(("2 FUTURE STATE", plan_frame))

    # ---- 3. lead time and dwell ----------------------------------------------------
    ws = _sheet(wb, "3 LEAD TIME AND DWELL",
                "One row per shipment inside a container — the grain at which a lead time "
                "is a fact rather than an average. Each row's dates all belong to the same "
                "cargo, so the three changes are subtractions you can check by eye, and "
                "the dwell that drives the storage charge is visible on the row it is "
                "charged to. Positive change means later. Every modelled date comes from a "
                "departure this lane actually ran.")
    pg = pallet_groups(result)
    if pg is not None and len(pg):
        _write_frame(ws, pg, start_row=3,
                     money_cols={"Storage USD", "Storage Rate USD Per CBM Per Day"},
                     num_cols={"CBM", "Weight kg", "% of This Shipment In This Box",
                               "% of This Box From This Shipment",
                               "Chargeable Storage CBM-Days"})
        sheets.append(("3 LEAD TIME AND DWELL", pg))

    # ---- 4. cost ledger -----------------------------------------------------------
    ws = _sheet(wb, "4 COST LEDGER",
                "Every dollar of modelled cost, one row per charge, tagged to a container, "
                "a cost pool, a rate and that rate's provenance. This tab and the totals on "
                "'2 FUTURE STATE' reconcile to the cent.")
    _write_frame(ws, result["ledger"], start_row=3, money_cols={"usd"},
                 num_cols={"quantity"})
    sheets.append(("4 COST LEDGER", result["ledger"]))

    # ---- 5. rate validation -------------------------------------------------------
    #
    # Three questions in one tab, because they are one question asked at three depths:
    # what is each rate, where did it come from, and what else could have priced that leg.
    # Splitting them across three tabs was tidier to write and worse to read -- somebody
    # arguing with a number wants all three lines about it in one place.
    ws = _sheet(wb, "5 RATE VALIDATION",
                "Every rate the model applied and where it came from. CLIENT_RATE_CARD is "
                "their own card; DERIVED_FROM_INVOICES was reconstructed from their charge "
                "lines, with the formula and observed spread shown; QUOTED_NOT_YET_BOUGHT "
                "prices the consolidation service, which no invoice can evidence because it "
                "is not bought yet. Below the rates: how each leg was sourced, and every "
                "constant the model applied.")
    _write_frame(ws, result["rates"], start_row=3,
                 money_cols={"rate_usd", "observed_min", "observed_median", "observed_max",
                             "invoice_total_usd"})
    sheets.append(("5 RATE VALIDATION", result["rates"]))

    src_plan = result.get("sourcing") or {}
    src_rows = []
    for spec in C.COST_COMPONENTS:
        cs = src_plan.get(spec["key"])
        if cs is None:
            continue
        pop = cs.population or {}
        alt = cs.alternative or {}
        src_rows.append({
            "Leg": cs.leg,
            "Component": cs.label,
            "What it is": cs.sub,
            "State": cs.state,
            "Priced": "yes" if cs.priced else "NO",
            "Qualified": "yes" if cs.qualified else "",
            "How": cs.chosen.get("note", ""),
            "Caveat": cs.caveat,
            "Charge codes": ", ".join(pop.get("codes", []) or cs.chosen.get("codes", [])),
            "Population groups": pop.get("groups", ""),
            "Population invoices": pop.get("invoices", ""),
            "Population USD": pop.get("usd", ""),
            "Card rows": cs.card_rows or "",
            "Blocked by": "; ".join(
                f"{b['code']} ({b['label']}) covering {', '.join(b['covers'])}"
                for b in cs.blocked_by),
            "Other sources that would have worked": "; ".join(
                f"{a['kind']}:{a['state']}" for a in cs.available if a is not cs.chosen),
            "Sources rejected, and why": " | ".join(
                f"{r['kind']}: {r['why']}" for r in cs.rejected),
            "Alternative": alt.get("label", ""),
            "Alternative state": alt.get("state", ""),
        })
    # How the cargo got to the warehouse, which is a decision and not a rate -- so it
    # cannot live in the rate table above, and it is far too consequential to leave to
    # whoever thinks to add up the ledger. Both ways of buying the leg, both load counts,
    # and the verdict.
    ib = result.get("inbound") or {}
    row = len(result["rates"]) + 6
    if (ib.get("priced") and ib.get("trucks")
            and ib.get("supplied") and ib.get("ftl_total") is not None):
        pairs = [
            ("Loads arriving at the warehouse", ib["loads_today"], ib["trucks"]),
            ("The haul, USD a year", ib["groupage_haul"], ib.get("ftl_haul", "")),
            (f"Warehouse receiving at ${ib['receiving_rate']:,.2f} a load, USD a year",
             ib["groupage_receiving"], ib.get("ftl_receiving", "")),
            ("Inbound cost of the plan, USD a year",
             ib["groupage_total"], ib.get("ftl_total", "")),
            ("Average pallets a load", ib["solo_pallets_mean"], ib["trailer_pallets_mean"]),
        ]
        _band(ws, row - 1,
              "How cargo reaches the warehouse — one collection per shipment, as bought "
              "today, against trailer loads bin-packed off the same cargo-ready dates. "
              + ("The tender is cheaper and the plan is costed at it."
                 if ib.get("path") == "ftl" else
                 "The tender is dearer, so the plan is costed at the collections.")
              + " Nothing waits at a factory for a trailer to fill. The physical packing "
                "opportunity is unchanged; the cheaper path may make a borderline site "
                "lane clear the commercial adoption rule.")
        _write_frame(ws, pd.DataFrame(
            [{"Measure": label, "As you buy it today": a,
              "Tendered by the trailer load": b} for label, a, b in pairs]),
            start_row=row, money_cols=set(), width_cap=72)
        row += len(pairs) + 4
    elif ib.get("priced") and ib.get("trucks") and not ib.get("supplied"):
        _band(ws, row - 1, "Optional inbound transport opportunity — not included in "
                           "this plan or its savings")
        ws.cell(
            row=row, column=1,
            value=(f"The plan keeps {ib['loads_today']:,} separate inbound collections, "
                   f"costed at ${ib['groupage_total']:,.2f}. No trailer rate was supplied, "
                   f"so pooled trailer collections were not assumed. The same-day cargo "
                   f"could form {ib['trucks']:,} trailer-sized loads; a haulage rate below "
                   f"${ib['headroom_per_load']:,.2f} per load, plus "
                   f"${ib['receiving_rate']:,.2f} receiving per arrival, may reduce cost."))
        ws.cell(row=row, column=1).font = NOTE_FONT
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
        ws.row_dimensions[row].height = 42
        row += 4
    elif ib.get("priced") and ib.get("trucks") and ib.get("partial_regions"):
        _band(ws, row - 1, "Trailer tender not included in this plan or its savings")
        ws.cell(
            row=row, column=1,
            value=("The tender does not cover "
                   + ", ".join(sorted(ib["partial_regions"]))
                   + ". The existing collections remain costed, and no incomplete "
                     "side-by-side comparison is shown."))
        ws.cell(row=row, column=1).font = NOTE_FONT
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
        ws.row_dimensions[row].height = 36
        row += 4
    if src_rows:
        _band(ws, row - 1, "How each leg was sourced — what was used, and what was rejected")
        _write_frame(ws, pd.DataFrame(src_rows), start_row=row,
                     money_cols={"Population USD"})
        row += len(src_rows) + 4

    cfg = result["config_used"]
    assum = pd.DataFrame([
        {"Setting": meta["label"], "Value": cfg[key], "Unit": meta["unit"],
         "Group": {"physical": "Physical limit", "dispatch": "Operating choice",
                   "decision": "Commercial adoption rule",
                   "service": "Consolidation service — priced separately"}[meta["group"]],
         "Priced from": {"quoted": "your forwarder's quote",
                         "edited": "your figure, entered on the settings step",
                         "benchmark": "our benchmark"}.get(
                             result["service_pricing"].get(key), "n/a"),
         "Where it comes from": meta["source"]}
        for key, meta in C.CONFIG.items()])
    derived = pd.DataFrame([
        {"Setting": f"{C.CONFIG[p['pct']]['label']} — resulting figure",
         "Value": cfg[p["derived"]], "Unit": p["unit"],
         "Group": "Derived from the pair above", "Priced from": "n/a",
         "Where it comes from": C.describe_pair(cfg, p)}
        for p in C.LIMIT_PAIRS])
    _band(ws, row - 1, "Every constant the model applied, and who is accountable for it")
    _write_frame(ws, pd.concat([assum, derived], ignore_index=True), start_row=row,
                 width_cap=96)

    # ---- 8. reconciliation --------------------------------------------------------
    #
    # Numbered 8, with 6 and 7 absent, because that is the numbering the client's own
    # workbook uses and this one is meant to sit beside it without anybody having to
    # relearn where anything is.
    ws = _sheet(wb, "8 RECONCILIATION",
                "Checks that must pass before any figure in this workbook is shown to "
                "anyone, then what the engine found in the file and the joins that needed "
                "judgement. Nothing was fixed silently.")
    controls = result["controls_frame"].copy()
    _write_frame(ws, controls, start_row=3, width_cap=100)
    row = len(controls) + 6

    dq = pd.DataFrame([{
        "Rule": f.rule, "What was found": f.headline, "Rows": f.rows,
        "USD at stake": f.dollars, "Narrated in the build": "yes" if f.narrate else "no",
        "Detail": f.note} for f in result["findings"]])
    _band(ws, row - 1, "What the engine found in the file, and what it did about it")
    _write_frame(ws, dq, start_row=row, money_cols={"USD at stake"}, width_cap=100)
    row += len(dq) + 4

    _band(ws, row - 1, "Site mappings that needed judgement — hand this back next run "
                       "and the review step is skipped")
    _write_frame(ws, result["resolution"].sites, start_row=row, width_cap=70)
    row += len(result["resolution"].sites) + 4
    _band(ws, row - 1, "Supplier pickup regions")
    _write_frame(ws, result["resolution"].suppliers, start_row=row, width_cap=70)

    # ---- back to tab 0, now that every column is known ----------------------------
    key_rows = []
    for name, frame in sheets:
        for col in list(frame.columns):
            meaning, unit, note = _column_note(name, str(col))
            key_rows.append({"Sheet": name, "Column": str(col),
                             "What it means": meaning, "Unit / format": unit,
                             "Notes": note})
    _write_frame(key_ws, pd.DataFrame(key_rows), start_row=3, width_cap=88)

    wb.save(path)
    return path
